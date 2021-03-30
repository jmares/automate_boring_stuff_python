import openpyxl, sys

# assumes an excel document with this name exists
wb = openpyxl.load_workbook('cell_inverter.xlsx')
# and that it contains one sheet with some data
sheet = wb.active

data = {}
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        data[(row, col)] = sheet.cell(row=row, column=col).value

# add new sheet for the inverted data
inv_sheet = wb.create_sheet()

# add inverted data
for key, val in data.items():
    col, row = key
    inv_sheet.cell(row=row, column=col).value = val

wb.save('cell_inverter.xlsx')