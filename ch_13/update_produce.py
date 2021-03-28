import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

# Loop through the rows and update the prices

for i in range(2, sheet.max_row):
    produce_name = sheet.cell(row=i, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=i, column=2).value = PRICE_UPDATES[produce_name]
    
wb.save('updatedProduceSales.xlsx')