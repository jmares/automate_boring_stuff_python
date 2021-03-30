import openpyxl, sys
from openpyxl.styles import Font

try:
    number = int(sys.argv[1])
    print(number)
except:
    print('Run the script with a number: eg. python multiplication_table.py 5')
    sys.exit()

wb = openpyxl.Workbook()
sheet = wb.active
bold = Font(bold=True)

# title row and column
for i in range(1,number+1):
    sheet.cell(row=1, column=i+1).value = i
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=1, column=i+1).font = bold
    sheet.cell(row=i+1, column=1).font = bold

# multiplication table

for row in range(1, number+1):
    for col in range(1, number+1):
            sheet.cell(row=row+1, column=col+1).value = row * col


wb.save('multiplication.xlsx')